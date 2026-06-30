"""
sincronizar_excel_bmax.py
Sincronização automática: BMAX CRITERIOS.xlsx (SharePoint) → Supabase.

Fluxo:
  1. Baixa o Excel do SharePoint via MSAL / Graph API
  2. Extrai revendas Ouro e Prata (aba REVENDA, col 0=Cliente, col 14=Nota, col 15=Classificação)
  3. Compara com Supabase (comercial_revendas_bmax)
  4. Adiciona novas revendas encontradas no Excel (com geocodificação automática)
  5. Atualiza rep/classe de revendas existentes se mudaram no Excel
  6. NUNCA altera o campo `ativo` — gerenciado exclusivamente pelo portal BMax
  7. NÃO toca lat/lng de revendas já geocodificadas

Env vars necessárias (GitHub Secrets ou .env):
  MSAL_TOKEN_CACHE       — cache JSON do token SharePoint (via renovar_msal_ci.py)
  SUPABASE_URL           — https://xxx.supabase.co
  SUPABASE_SERVICE_KEY   — service_role key (nunca no frontend)
"""

import os, sys, json, time, tempfile, io, requests
from pathlib import Path

# ── codificação segura no Windows ─────────────────────────────
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── carrega .env se existir ────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass

SB_URL = os.environ["SUPABASE_URL"]
SB_KEY = os.environ["SUPABASE_SERVICE_KEY"]

TENANT_ID  = "c0bbec8e-9949-4dcf-80b3-bd21689c33e4"
CLIENT_ID  = "d1fb2af8-a56b-41a0-8401-e87591360016"
SCOPE      = ["https://graph.microsoft.com/Sites.Read.All"]
DRIVE_ID   = "b!6Kvw0SoDTUiClr8MvMu7NuwGBC9zdONMq_ED-vRKE3W43Ls7sVoDTZSN_i_i5kpD"
EXCEL_ITEM = "015WUCIWDTZYXSMZAJGJGIUOY4XZKPVHD6"

# Col indices na aba REVENDA (linha 2 = cabeçalho, dados a partir linha 3)
COL_CLIENTE    = 0
COL_REP        = 1
COL_CIDADE_UF  = 2   # "Cidade/UF" ou "Cidade / UF"
COL_ESTADO     = 3
COL_NOTA       = 14  # nota numérica
COL_CLASS      = 15  # "Ouro", "Prata", "Não Aplica"


# ── Supabase helpers ───────────────────────────────────────────
def sb(path, method="GET", body=None, params=None):
    hdrs = {
        "apikey": SB_KEY,
        "Authorization": f"Bearer {SB_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation" if method == "POST"
                  else "return=minimal" if method in ("PATCH", "PUT")
                  else "count=exact",
    }
    url  = SB_URL.rstrip("/") + "/rest/v1" + path
    resp = requests.request(method, url, headers=hdrs,
                            json=body, params=params, timeout=20)
    if not resp.ok:
        raise RuntimeError(f"Supabase {method} {path} → {resp.status_code}: {resp.text[:300]}")
    if method in ("GET", "POST"):
        return resp.json()
    return resp


# ── MSAL / Graph API ───────────────────────────────────────────
def get_graph_token() -> str:
    import msal
    cache_env = os.environ.get("MSAL_TOKEN_CACHE", "")
    cache = msal.SerializableTokenCache()
    if cache_env:
        cache.deserialize(cache_env)
    else:
        # fallback: arquivo local (dev)
        local = Path(__file__).parent.parent.parent.parent.parent.parent / \
                "App Vendas/bsis_v3/backend/data/.oauth_token_cache.json"
        if local.exists():
            cache.deserialize(local.read_text(encoding="utf-8"))
    app = msal.PublicClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        token_cache=cache,
    )
    accounts = app.get_accounts()
    if not accounts:
        raise RuntimeError("MSAL: nenhuma conta em cache. Renove o token com renovar_msal_ci.py")
    result = app.acquire_token_silent(SCOPE, account=accounts[0])
    if not result or "access_token" not in result:
        raise RuntimeError(f"MSAL: token expirado — {result.get('error_description','')}")
    return result["access_token"]


def baixar_excel(token: str) -> bytes:
    hdrs = {"Authorization": f"Bearer {token}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{EXCEL_ITEM}/content",
        headers=hdrs, timeout=30
    )
    r.raise_for_status()
    return r.content


# ── parser Excel ───────────────────────────────────────────────
def parsear_excel(data: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    # Procura aba REVENDA
    aba = None
    for name in wb.sheetnames:
        if "REVENDA" in name.upper():
            aba = wb[name]
            break
    if not aba:
        raise RuntimeError(f"Aba REVENDA não encontrada. Abas disponíveis: {wb.sheetnames}")

    revendas = []
    rows = list(aba.iter_rows(values_only=True))
    # Cabeçalho na linha 2 (índice 1), dados a partir da linha 3 (índice 2)
    for row in rows[2:]:
        if not row or not row[COL_CLIENTE]:
            continue
        nome  = str(row[COL_CLIENTE]).strip()
        if not nome or nome.upper() in ("CLIENTE", "NOME", ""):
            continue
        classe_raw = str(row[COL_CLASS]).strip() if row[COL_CLASS] else ""
        if classe_raw not in ("Ouro", "Prata"):
            continue   # exclui Não Aplica e vazios

        rep_raw = str(row[COL_REP]).strip() if row[COL_REP] else ""
        estado  = str(row[COL_ESTADO]).strip() if row[COL_ESTADO] else ""

        # Extrai cidade de "Cidade/UF" ou da coluna própria
        cidade = ""
        if row[COL_CIDADE_UF]:
            parts = str(row[COL_CIDADE_UF]).split("/")
            cidade = parts[0].strip()
            if not estado and len(parts) > 1:
                estado = parts[1].strip()

        revendas.append({
            "nome":    nome,
            "rep":     rep_raw or None,
            "cidade":  cidade or None,
            "estado":  estado or None,
            "classe":  classe_raw,
        })

    return revendas


# ── Geocodificação ─────────────────────────────────────────────
def geocodificar(cidade: str, estado: str) -> tuple[float, float] | None:
    time.sleep(1.2)
    # Usa o primeiro nome de cidades compostas (ex: "São Bento do Sul e Canoinhas")
    cidade_p = cidade.split(" e ")[0].strip()
    try:
        r = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": f"{cidade_p}, {estado}, Brasil",
                    "format": "json", "limit": 1, "countrycodes": "br"},
            headers={"User-Agent": "BMaxMotorPCI/1.0 (boxersoldas.com.br)"},
            timeout=12,
        )
        if r.ok and r.json():
            item = r.json()[0]
            return float(item["lat"]), float(item["lon"])
    except Exception as e:
        print(f"   !! Nominatim: {e}")
    return None


# ── main ───────────────────────────────────────────────────────
def main():
    print("=== BMax — Sincronização Excel → Supabase ===\n")

    # 1. Baixar Excel
    print("1. Autenticando no SharePoint...")
    token = get_graph_token()
    print("   Token obtido.")
    print("2. Baixando BMAX CRITERIOS.xlsx...")
    excel_bytes = baixar_excel(token)
    print(f"   Arquivo baixado ({len(excel_bytes):,} bytes)")

    # 2. Parsear
    print("3. Processando revendas do Excel...")
    excel_revs = parsear_excel(excel_bytes)
    print(f"   {len(excel_revs)} revendas Ouro/Prata encontradas")

    # 3. Carregar estado atual do Supabase
    print("4. Carregando Supabase...")
    sb_revs = sb("/comercial_revendas_bmax?select=id,nome,rep,cidade,estado,classe,lat,lng,ativo")
    print(f"   {len(sb_revs)} registros no banco")

    # Índice por nome normalizado
    def norm(s: str) -> str:
        return s.strip().upper() if s else ""

    sb_idx = {norm(r["nome"]): r for r in sb_revs}
    excel_nomes = {norm(r["nome"]) for r in excel_revs}

    adicionadas = inseridas = atualizadas = 0
    erros = []

    # 4. Processar cada revenda do Excel
    print("\n5. Sincronizando...\n")
    for rev in excel_revs:
        key = norm(rev["nome"])
        existing = sb_idx.get(key)

        if existing:
            updates = {}
            # Atualiza campos que mudaram (rep, classe, cidade, estado)
            for campo in ("rep", "classe", "cidade", "estado"):
                val_ex = str(existing.get(campo) or "").strip()
                val_xl = str(rev.get(campo) or "").strip()
                if val_xl and val_xl != val_ex:
                    updates[campo] = rev[campo]
            if updates:
                try:
                    sb(f"/comercial_revendas_bmax?id=eq.{existing['id']}", "PATCH", updates)
                    if "ativo" not in updates:  # não contar duplo
                        atualizadas += 1
                        print(f"  ↻ Atualizada: {rev['nome']} {list(updates.keys())}")
                except Exception as e:
                    erros.append(f"{rev['nome']}: {e}")
        else:
            # Nova revenda — geocodificar
            lat, lng = None, None
            if rev.get("cidade") and rev.get("estado"):
                coords = geocodificar(rev["cidade"], rev["estado"])
                if coords:
                    lat, lng = coords
                    print(f"  ✚ Nova:       {rev['nome']} — {rev['cidade']}/{rev['estado']} → {lat:.4f},{lng:.4f}")
                else:
                    print(f"  ✚ Nova:       {rev['nome']} — sem geocode")
            else:
                print(f"  ✚ Nova:       {rev['nome']} — sem cidade/estado")
            registro = {**rev, "lat": lat, "lng": lng, "ativo": True}
            try:
                sb("/comercial_revendas_bmax", "POST", registro)
                inseridas += 1
                adicionadas += 1
            except Exception as e:
                erros.append(f"{rev['nome']}: {e}")

    print(f"\n=== Resultado ===")
    print(f"  Inseridas:   {inseridas}")
    print(f"  Atualizadas: {atualizadas}")
    print(f"  (ativo gerenciado pelo portal — sem inativação automática)")
    if erros:
        print(f"  Erros ({len(erros)}):")
        for e in erros:
            print(f"    • {e}")
        sys.exit(1)
    else:
        print("  OK — sem erros")


if __name__ == "__main__":
    main()
